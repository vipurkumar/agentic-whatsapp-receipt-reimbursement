"""Excel logger with query and delete capabilities using openpyxl."""

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from receipt_processor import ReceiptData

HEADERS = ["#", "Date", "Merchant", "Expense Type", "Amount", "Currency", "Description", "Logged At", "Image Path"]
DEFAULT_PATH = "reimbursements.xlsx"


def _get_or_create_workbook(file_path: str = DEFAULT_PATH) -> tuple[Workbook, bool]:
    """Open existing workbook or create a new one with styled headers."""
    path = Path(file_path)
    if path.exists():
        return load_workbook(file_path), False

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Reimbursements"

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")

    wb.save(file_path)
    return wb, True


def is_duplicate(data: ReceiptData, file_path: str = DEFAULT_PATH) -> bool:
    """Check if a receipt with the same date, merchant, and amount already exists."""
    path = Path(file_path)
    if not path.exists():
        return False

    wb = load_workbook(file_path)
    ws = wb.active
    assert ws is not None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == data.date and row[2] == data.merchant and str(row[4]) == str(data.amount):
            return True
    return False


def append_receipt(data: ReceiptData, file_path: str = DEFAULT_PATH, image_path: str = "") -> int:
    """Append receipt data to Excel file. Returns the sequence number."""
    wb, _ = _get_or_create_workbook(file_path)
    ws = wb.active
    assert ws is not None

    seq_number = ws.max_row  # max_row includes header, so this equals next sequence
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    row = [
        seq_number,
        data.date,
        data.merchant,
        data.expense_type,
        data.amount,
        data.currency,
        data.description,
        timestamp,
        image_path,
    ]

    ws.append(row)
    wb.save(file_path)
    return seq_number


def query_receipts(filter_type: str, filter_value: str | None = None, file_path: str = DEFAULT_PATH) -> list[dict]:
    """Query receipts with various filters.

    filter_type: 'all', 'month', 'expense_type', 'last_n', 'merchant', 'summary'
    filter_value: value for the filter (e.g. '03/2026', 'Meals', '5', 'Amazon')
    """
    path = Path(file_path)
    if not path.exists():
        return []

    wb = load_workbook(file_path)
    ws = wb.active
    assert ws is not None

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "seq_number": row[0],
            "date": row[1],
            "merchant": row[2],
            "expense_type": row[3],
            "amount": str(row[4]),
            "currency": str(row[5]),
            "description": row[6],
            "logged_at": row[7],
            "image_path": row[8] if len(row) > 8 else "",
        })

    if filter_type == "all":
        return rows

    if filter_type == "month" and filter_value:
        # filter_value expected as "MM/YYYY"
        return [r for r in rows if r["date"] and r["date"].endswith(filter_value)]

    if filter_type == "expense_type" and filter_value:
        return [r for r in rows if r["expense_type"] and r["expense_type"].lower() == filter_value.lower()]

    if filter_type == "last_n" and filter_value:
        n = int(filter_value)
        return rows[-n:]

    if filter_type == "merchant" and filter_value:
        return [r for r in rows if r["merchant"] and filter_value.lower() in r["merchant"].lower()]

    if filter_type == "summary":
        return rows  # Agent will compute the summary from raw data

    return rows


def delete_receipt(seq_number: int, file_path: str = DEFAULT_PATH) -> bool:
    """Delete a receipt by sequence number. Re-numbers remaining rows."""
    path = Path(file_path)
    if not path.exists():
        return False

    wb = load_workbook(file_path)
    ws = wb.active
    assert ws is not None

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is not None and int(cell_val) == seq_number:  # type: ignore[arg-type]
            target_row = row_idx
            break

    if target_row is None:
        return False

    ws.delete_rows(target_row)

    # Re-number remaining rows
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)

    wb.save(file_path)
    return True


def get_summary(file_path: str = DEFAULT_PATH) -> dict:
    """Get summary totals by currency and expense type."""
    rows = query_receipts("all", file_path=file_path)

    by_currency: dict[str, float] = {}
    by_type: dict[str, float] = {}

    for r in rows:
        amount = float(r["amount"])
        currency = r["currency"]
        expense_type = r["expense_type"]

        by_currency[currency] = by_currency.get(currency, 0.0) + amount
        key = f"{expense_type} ({currency})"
        by_type[key] = by_type.get(key, 0.0) + amount

    return {
        "total_receipts": len(rows),
        "by_currency": {k: round(v, 2) for k, v in by_currency.items()},
        "by_expense_type": {k: round(v, 2) for k, v in by_type.items()},
    }
