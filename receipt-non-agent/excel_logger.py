"""Fallback Excel logger using openpyxl (not wired into main pipeline)."""

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from receipt_processor import ReceiptData

HEADERS = ["#", "Date", "Merchant", "Expense Type", "Amount", "Currency", "Description", "Logged At"]
DEFAULT_PATH = "reimbursements.xlsx"


def _get_or_create_workbook(file_path: str = DEFAULT_PATH) -> tuple[Workbook, bool]:
    """Open existing workbook or create a new one with styled headers."""
    path = Path(file_path)
    if path.exists():
        return load_workbook(file_path), False

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Reimbursements"

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")

    wb.save(file_path)
    return wb, True


def is_duplicate(data: ReceiptData, file_path: str = DEFAULT_PATH) -> bool:
    """Check if a receipt with the same date, merchant, and amount already exists."""
    path = Path(file_path)
    if not path.exists():
        return False

    wb = load_workbook(file_path)
    ws = wb.active
    assert ws is not None

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: #, Date, Merchant, Expense Type, Amount, Currency
        if row[1] == data.date and row[2] == data.merchant and str(row[4]) == str(data.amount):
            return True
    return False


def append_receipt(data: ReceiptData, file_path: str = DEFAULT_PATH) -> int:
    """Append receipt data to Excel file. Returns the sequence number."""
    wb, _ = _get_or_create_workbook(file_path)
    ws = wb.active
    assert ws is not None

    seq_number = ws.max_row  # max_row includes header, so this equals next sequence
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    row = [
        seq_number,
        data.date,
        data.merchant,
        data.expense_type,
        data.amount,
        data.currency,
        data.description,
        timestamp,
    ]

    ws.append(row)
    wb.save(file_path)
    return seq_number
